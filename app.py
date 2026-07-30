# Mengimpor library yang dibutuhkan
from flask import Flask, render_template, jsonify, request
import sqlite3  # Database SQLite
from openpyxl import Workbook, load_workbook  # Membuat dan membaca file Excel
from openpyxl.styles import Font, PatternFill, Alignment  # Styling Excel
import os  # Mengelola file dan folder

# Membuat aplikasi Flask
app = Flask(__name__)

# ===============================
# KONEKSI DATABASE SQLITE
# ===============================
def get_db():
    # Menghubungkan ke database.db
    conn = sqlite3.connect("database.db")

    # Agar hasil query bisa dipanggil berdasarkan nama kolom
    conn.row_factory = sqlite3.Row

    return conn

# ===============================
# HALAMAN UTAMA
# ===============================
@app.route("/")
def home():

    # Menampilkan halaman index.html
    return render_template("index.html")

# ===============================
# AMBIL DATA PRODUK
# ===============================
@app.route("/produk")
def produk():

    db = get_db()

    db.execute("""
        CREATE TABLE IF NOT EXISTS produk (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT,
            harga TEXT
        )
    """)

    data = db.execute("SELECT * FROM produk").fetchall()

    db.commit()
    db.close()

    return jsonify([dict(row) for row in data])


# ===============================
# CEK DATA PESANAN SQLITE
# ===============================
@app.route("/cek")
def cek():

    db = get_db()

    try:
        data = db.execute("""
            SELECT *
            FROM pesanan
            ORDER BY id DESC
        """).fetchall()

        return jsonify([dict(row) for row in data])

    except Exception as e:

        return jsonify({
            "status": "error",
            "message": str(e)
        })

    finally:
        db.close()


# ===============================
# SIMPAN PESANAN
# ===============================
@app.route("/pesan", methods=["POST"])
def simpan_pesanan():

    try:

        data = request.get_json()

        print("=" * 50)
        print("DATA MASUK DARI FORM")
        print(data)
        print("=" * 50)

        nama = data.get("nama")
        jenis = data.get("jenis")
        alamat = data.get("alamat")
        nohp = data.get("nohp")
        metode = data.get("metode")
        hari = data.get("hari")
        tanggal = data.get("tanggal")
        jam = data.get("jam")
        harga = data.get("harga")
        request_tambahan = data.get("request")

        # ======================
        # SQLITE
        # ======================

        db = get_db()

        db.execute("""
            CREATE TABLE IF NOT EXISTS pesanan (

                id INTEGER PRIMARY KEY AUTOINCREMENT,

                nama TEXT,
                jenis TEXT,
                alamat TEXT,
                nohp TEXT,
                metode TEXT,
                hari TEXT,
                tanggal TEXT,
                jam TEXT,
                harga TEXT,
                request_tambahan TEXT
            )
        """)

        db.execute("""
            INSERT INTO pesanan
            (
                nama,
                jenis,
                alamat,
                nohp,
                metode,
                hari,
                tanggal,
                jam,
                harga,
                request_tambahan
            )

            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            nama,
            jenis,
            alamat,
            nohp,
            metode,
            hari,
            tanggal,
            jam,
            harga,
            request_tambahan
        ))

        db.commit()

        print("DATA BERHASIL MASUK SQLITE")

        db.close()

        # ======================
        # EXCEL
        # ======================

        file = os.path.abspath("pesanan.xlsx")

        print("LOKASI FILE EXCEL:")
        print(file)

        if not os.path.exists(file):

            print("MEMBUAT FILE EXCEL BARU")

            wb = Workbook()
            ws = wb.active

            ws.title = "Data Pesanan"

            headers = [
                "Nama",
                "Jenis Pesanan",
                "Alamat",
                "No HP",
                "Metode",
                "Hari",
                "Tanggal",
                "Jam",
                "Harga",
                "Request Tambahan"
            ]

            ws.append(headers)

            for cell in ws[1]:

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    start_color="D63384",
                    end_color="D63384",
                    fill_type="solid"
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 18
            ws.column_dimensions['J'].width = 35

        else:

            print("MEMBUKA FILE EXCEL YANG SUDAH ADA")

            wb = load_workbook(file)
            ws = wb.active

        print("BARIS SEBELUM TAMBAH :", ws.max_row)

        ws.append([
            nama,
            jenis,
            alamat,
            nohp,
            metode,
            hari,
            tanggal,
            jam,
            harga,
            request_tambahan
        ])

        print("BARIS SESUDAH TAMBAH :", ws.max_row)

        wb.save(file)
        wb.close()

        print("DATA BERHASIL DISIMPAN KE EXCEL")

        return jsonify({
            "status": "success",
            "message": "Pesanan berhasil disimpan"
        })

    except Exception as e:

        print("ERROR:", str(e))

        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# ===============================
# MENJALANKAN SERVER
# ===============================
if __name__ == "__main__":

    # Menjalankan Flask dalam mode debug
    app.run(debug=True)